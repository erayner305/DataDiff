import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Read Excel files into dataframes
print("Ensure you follow the readme before continuing")
file1 = input("Input the sanitas data file's name without extension: ")
file2 = input("Input the comparison data file's name without extension: ")

df1 = pd.read_excel(f"{file1}.xlsx")
df2 = pd.read_excel(f"{file2}.xlsx")

df1['Date'] = df1['Date'].dt.date
df2['Date'] = df2['Date'].dt.date

# Columns to identify unique rows
key_columns = ['Well', 'Date', 'Constituent']

# Columns to check for differences
diff_columns = ['MDL', 'PQL', 'Flags', 'Obs']

# Merge the dataframes based on the key columns
merged_df = pd.merge(df1, df2, on=key_columns, how='outer', indicator=True, suffixes=(f'_{file1}', f'_{file2}'))

# Identify rows unique to each file
only_in_df1 = merged_df[merged_df['_merge'] == 'left_only'][key_columns + [f"{col}_{file1}" for col in diff_columns]]
only_in_df2 = merged_df[merged_df['_merge'] == 'right_only'][key_columns + [f"{col}_{file2}" for col in diff_columns]]

# Identify rows that exist in both but have differing values in specified columns
common_rows = merged_df[merged_df['_merge'] == 'both']
def check_differences(row):
    for col in diff_columns:
        val1 = row[f"{col}_{file1}"]
        val2 = row[f"{col}_{file2}"]

        if col == 'Observed':
            # Skip the check if either value starts with '<'
            if isinstance(val1, str) and val1.startswith('<'):
                continue
            if isinstance(val2, str) and val2.startswith('<'):
                continue

        if val1 != val2:
            return True

    return False

differing_rows = common_rows[common_rows.apply(check_differences, axis=1)]
# Output results to Excel files
only_in_df1.to_excel(f"only_in_{file1}.xlsx", index=False)
only_in_df2.to_excel(f"only_in_{file2}.xlsx", index=False)

# Define a function that identifies which cells differ in a given row
def highlight_differences(row, worksheet, row_num):
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    start_col = len(key_columns) + 1  # Assuming key columns are in the beginning

    for col_idx, col in enumerate(diff_columns):
        val1 = row[f"{col}_{file1}"]
        val2 = row[f"{col}_{file2}"]

        if val1 != val2:
            # Highlight the cell. Note that Excel is 1-indexed
            # Column index starts after key columns
            worksheet.cell(row=row_num, column=start_col + col_idx).fill = yellow_fill


# Save differing_rows to Excel and open the workbook and worksheet
differing_rows.to_excel("differing_rows.xlsx", index=False)
workbook = Workbook()
workbook = openpyxl.load_workbook("differing_rows.xlsx")
worksheet = workbook.active

# Apply highlighting
for row_idx, row in differing_rows.iterrows():
    highlight_differences(row, worksheet, row_idx + 2)  # row_idx + 2 because Excel is 1-indexed and we have a header

# Save the changes
workbook.save("differing_rows.xlsx")

print(f"Unique and differing rows have been saved to 'only_in_{file1}.xlsx', 'only_in_{file2}.xlsx', and 'differing_rows.xlsx'")
